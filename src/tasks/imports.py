import asyncio
import csv
import os
import sys
import tempfile
import uuid

from openpyxl import load_workbook

from src.api.v1.schemas.batch import BatchCreate
from src.celery_app import celery_app
from src.core.database import AsyncSessionLocal
from src.data.repositories.batch_repository import BatchRepository
from src.data.repositories.webhook_repository import WebhookRepository
from src.data.repositories.work_center_repository import WorkCenterRepository
from src.domain.exceptions import BatchAlreadyExistsError
from src.domain.services.batch_service import BatchService
from src.domain.services.webhook_service import WebhookService
from src.storage.minio_service import MinIOService

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


async def _import_batches_from_file_async(object_name: str, user_id: int | None) -> dict:
    is_csv = object_name.lower().endswith(".csv")
    extension = "csv" if is_csv else "xlsx"
    local_path = os.path.join(tempfile.gettempdir(), f"import_{uuid.uuid4().hex[:8]}.{extension}")

    storage = MinIOService()
    storage.download_file(bucket="imports", object_name=object_name, file_path=local_path)

    try:
        if is_csv:
            with open(local_path, encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter=";")
                all_rows = list(reader)
            rows = all_rows[1:]  # первая строка — заголовок
        else:
            wb = load_workbook(local_path)
            sheet = wb.active
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
    finally:
        os.remove(local_path)

    created = 0
    errors = []

    async with AsyncSessionLocal() as session:
        service = BatchService(
            batch_repo=BatchRepository(session),
            work_center_repo=WorkCenterRepository(session),
            webhook_service=WebhookService(webhook_repo=WebhookRepository(session)),
        )

        for row_number, row in enumerate(rows, start=2):
            (
                batch_number, batch_date, nomenclature, work_center_name,
                work_center_identifier, shift, team, ekn_code,
                task_description, shift_start, shift_end,
            ) = row

            try:
                batch_data = BatchCreate(
                    batch_number=batch_number, batch_date=batch_date, nomenclature=nomenclature,
                    work_center_name=work_center_name, work_center_identifier=work_center_identifier,
                    shift=shift, team=team, ekn_code=ekn_code, task_description=task_description,
                    shift_start=shift_start, shift_end=shift_end,
                )
                await service.create_batch(batch_data)
                created += 1
            except BatchAlreadyExistsError:
                errors.append({"row": row_number, "error": "Duplicate batch number and date"})
            except (ValueError, TypeError) as e:
                errors.append({"row": row_number, "error": f"Invalid row data: {e}"})

        await session.commit()

        webhook_service = WebhookService(webhook_repo=WebhookRepository(session))
        await webhook_service.dispatch_event(
            "import_completed",
            {
                "total_rows": len(rows),
                "created": created,
                "skipped": len(errors),
                "errors": errors,
            },
        )
        await session.commit()

    return {
        "success": True,
        "total_rows": len(rows),
        "created": created,
        "skipped": len(errors),
        "errors": errors,
    }


@celery_app.task(bind=True, max_retries=1)
def import_batches_from_file(self, file_url: str, user_id: int | None = None) -> dict:
    return asyncio.run(_import_batches_from_file_async(file_url, user_id))
