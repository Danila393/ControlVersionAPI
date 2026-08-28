import asyncio
import os
import sys
import tempfile
import uuid
from datetime import UTC, datetime, timedelta

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from src.celery_app import celery_app
from src.core.database import AsyncSessionLocal
from src.data.repositories.batch_repository import BatchRepository
from src.data.repositories.webhook_repository import WebhookRepository
from src.domain.services.webhook_service import WebhookService
from src.storage.minio_service import MinIOService

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

# Встроенные шрифты reportlab (Helvetica и т.п.) не умеют в кириллицу — нужен
# TTF-шрифт со всеми буквами. Берём системный Arial (есть в Windows) —
# это дев-окружение; для реального деплоя (например, в Docker/Linux)
# шрифт нужно будет положить файлом прямо в проект, а не полагаться на ОС.
_CYRILLIC_FONT = "Arial"
_WINDOWS_ARIAL_PATH = "C:/Windows/Fonts/arial.ttf"
if _CYRILLIC_FONT not in pdfmetrics.getRegisteredFontNames() and os.path.exists(_WINDOWS_ARIAL_PATH):
    pdfmetrics.registerFont(TTFont(_CYRILLIC_FONT, _WINDOWS_ARIAL_PATH))


def _build_report_workbook(batch) -> Workbook:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Информация о партии"

    sheet.append(["Номер партии:", batch.batch_number])
    sheet.append(["Дата партии:", str(batch.batch_date)])
    sheet.append(["Статус:", "Закрыта" if batch.is_closed else "Открыта"])
    sheet.append(["Рабочий центр:", batch.work_center.name])
    sheet.append(["Смена:", batch.shift])
    sheet.append(["Бригада:", batch.team])
    sheet.append(["Номенклатура:", batch.nomenclature])
    sheet.append(["Начало смены:", str(batch.shift_start)])
    sheet.append(["Окончание смены:", str(batch.shift_end)])

    products_sheet = wb.create_sheet("Продукция")
    products_sheet.append(["ID", "Уникальный код", "Аггрегирована", "Дата аггрегации"])
    for product in batch.products:
        products_sheet.append([
            product.id,
            product.unique_code,
            "Да" if product.is_aggregated else "Нет",
            str(product.aggregated_at) if product.aggregated_at else "-",
        ])

    total, aggregated_count, avg_speed = _stats(batch)
    statistics_sheet = wb.create_sheet("Статистика")
    statistics_sheet.append(["Всего продукции:", total])
    statistics_sheet.append(["Аггрегировано:", aggregated_count])
    statistics_sheet.append(["Осталось", total - aggregated_count])
    statistics_sheet.append(["Процент выполнения", aggregated_count / total * 100 if total > 0 else 0])
    statistics_sheet.append(["Средняя скорость:", avg_speed])

    return wb


def _stats(batch) -> tuple[int, int, float]:
    total = len(batch.products)
    aggregated_count = sum(1 for p in batch.products if p.is_aggregated)
    shift_hours = (batch.shift_end - batch.shift_start).total_seconds() / 3600
    avg_speed = aggregated_count / shift_hours if shift_hours > 0 else 0
    return total, aggregated_count, avg_speed


def _build_report_pdf(batch, local_path: str) -> None:
    total, aggregated_count, avg_speed = _stats(batch)

    styles = getSampleStyleSheet()
    font = _CYRILLIC_FONT if _CYRILLIC_FONT in pdfmetrics.getRegisteredFontNames() else "Helvetica"
    heading = ParagraphStyle("heading", parent=styles["Heading1"], fontName=font)
    body = ParagraphStyle("body", parent=styles["Normal"], fontName=font)

    info_rows = [
        ["Номер партии:", str(batch.batch_number)],
        ["Дата партии:", str(batch.batch_date)],
        ["Статус:", "Закрыта" if batch.is_closed else "Открыта"],
        ["Рабочий центр:", batch.work_center.name],
        ["Смена:", batch.shift],
        ["Бригада:", batch.team],
        ["Номенклатура:", batch.nomenclature],
        ["Начало смены:", str(batch.shift_start)],
        ["Окончание смены:", str(batch.shift_end)],
    ]
    stats_rows = [
        ["Всего продукции:", str(total)],
        ["Аггрегировано:", str(aggregated_count)],
        ["Осталось:", str(total - aggregated_count)],
        ["Процент выполнения:", f"{aggregated_count / total * 100 if total > 0 else 0:.1f}%"],
        ["Средняя скорость:", f"{avg_speed:.2f} ед/час"],
    ]

    table_style = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ])

    doc = SimpleDocTemplate(local_path, pagesize=A4)
    elements = [
        Paragraph("Отчёт по партии", heading),
        Spacer(1, 12),
        Paragraph("Информация о партии", body),
        Table(info_rows, style=table_style),
        Spacer(1, 16),
        Paragraph("Статистика", body),
        Table(stats_rows, style=table_style),
    ]
    doc.build(elements)


async def _generate_batch_report_async(batch_id: int, format: str = "excel") -> dict:
    async with AsyncSessionLocal() as session:
        batch_repo = BatchRepository(session)
        batch = await batch_repo.get_by_id(batch_id)

        if batch is None:
            return {"success": False, "error": f"Batch {batch_id} not found"}

        is_pdf = format == "pdf"
        extension = "pdf" if is_pdf else "xlsx"
        file_name = f"batch_{batch_id}_report_{uuid.uuid4().hex[:8]}.{extension}"
        local_path = os.path.join(tempfile.gettempdir(), file_name)

        if is_pdf:
            _build_report_pdf(batch, local_path)
        else:
            _build_report_workbook(batch).save(local_path)

        try:
            file_size = os.path.getsize(local_path)
            expires_days = 7
            storage = MinIOService()
            file_url = storage.upload_file(
                bucket="reports",
                file_path=local_path,
                object_name=file_name,
                expires_days=expires_days,
            )
        finally:
            os.remove(local_path)

        webhook_service = WebhookService(webhook_repo=WebhookRepository(session))
        await webhook_service.dispatch_event(
            "report_generated",
            {
                "batch_id": batch_id,
                "report_type": format,
                "file_url": file_url,
                "expires_at": str(datetime.now(UTC) + timedelta(days=expires_days)),
            },
        )
        await session.commit()

    return {
        "success": True,
        "file_url": file_url,
        "file_name": file_name,
        "file_size": file_size,
        "expires_at": str(datetime.now(UTC) + timedelta(days=expires_days)),
    }


@celery_app.task(bind=True, max_retries=3)
def generate_batch_report(self, batch_id: int, format: str = "excel", user_email: str | None = None) -> dict:
    return asyncio.run(_generate_batch_report_async(batch_id, format))
